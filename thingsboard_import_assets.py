#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Skript zum Importieren von Assets in Thingsboard aus Excel-Datei.

Erstellt Assets mit Relations basierend auf der Hierarchie:
- JHRZO_0001 (existiert bereits) - Root Asset (Level 6)
  - Level 5 (Spalte C) → Asset Profil "City" → Relation zu Level 6 (JHRZO_0001)
    - Level 4 (Spalte D) → Asset Profil "Building" → Relation zu Level 5 Asset
      - Level 3 (Spalte E) → Asset Profil "Floor" → Relation zu Level 4 Asset
        - Level 2 (Spalte F) → Asset Profil "Area" → Relation zu Level 3 Asset
          - Level 1 (Spalte G) → Asset Profil "Room" → Relation zu Level 2 Asset

Excel-Struktur:
- Spalte A = Sensor DevEUI (wird ignoriert)
- Spalte B = Level 6 (Root, bereits importiert als JHRZO_0001)
- Spalte C = Level 5 (City)
- Spalte D = Level 4 (Building)
- Spalte E = Level 3 (Floor)
- Spalte F = Level 2 (Area)
- Spalte G = Level 1 (Room)
- Spalte H = Ebene 6 (Asset-Name Level 6, wird geschrieben)
- Spalte I = Ebene 5 (Asset-Name Level 5, wird geschrieben)
- Spalte J = Ebene 4 (Asset-Name Level 4, wird geschrieben)
- Spalte K = Ebene 3 (Asset-Name Level 3, wird geschrieben)
- Spalte L = Ebene 2 (Asset-Name Level 2, wird geschrieben)
- Spalte M = Ebene 1 (Asset-Name Level 1, wird geschrieben)

Assets werden durchnummeriert ab JHRZO_0002 mit Prefix JHRZO_xxxxx.
"""

import pandas as pd
import requests
import sys
import os
from dotenv import load_dotenv
from heatmanager_common.config import THINGSBOARD_BASE_URL, THINGSBOARD_USERNAME, THINGSBOARD_PASSWORD

# .env Datei laden
load_dotenv()

# Thingsboard API Token
token = None

def login_to_thingsboard():
    """Login zu Thingsboard und erhalte Token"""
    global token
    
    if token:
        return token
    
    try:
        print(f"🔐 Login zu Thingsboard: {THINGSBOARD_BASE_URL}")
        
        login_url = f"{THINGSBOARD_BASE_URL}/api/auth/login"
        login_data = {
            "username": THINGSBOARD_USERNAME,
            "password": THINGSBOARD_PASSWORD
        }
        
        response = requests.post(
            login_url,
            json=login_data,
            headers={'Content-Type': 'application/json'},
            timeout=30
        )
        
        if response.status_code == 200:
            token = response.json().get('token')
            print(f"✅ Login erfolgreich")
            return token
        else:
            print(f"❌ Login fehlgeschlagen: {response.status_code}")
            print(f"   Response: {response.text}")
            return None
            
    except Exception as e:
        print(f"❌ Fehler beim Login: {e}")
        import traceback
        traceback.print_exc()
        return None

def get_asset_by_name(asset_name, verbose=False):
    """Sucht ein Asset nach Namen"""
    try:
        url = f"{THINGSBOARD_BASE_URL}/api/tenant/assets"
        
        headers = {
            'X-Authorization': f'Bearer {token}',
            'Content-Type': 'application/json'
        }
        
        # Suche mit Pagination - hole alle Assets
        page = 0
        page_size = 100
        has_next = True
        
        while has_next:
            params = {
                'page': page,
                'pageSize': page_size,
                'textSearch': asset_name
            }
            
            response = requests.get(url, headers=headers, params=params, timeout=30)
            
            if response.status_code == 200:
                data = response.json()
                assets = data.get('data', [])
                
                if verbose:
                    print(f"    Durchsuche {len(assets)} Assets (Seite {page})...")
                
                # Suche nach exakter Namensübereinstimmung
                for asset in assets:
                    asset_name_found = asset.get('name', '')
                    if asset_name_found == asset_name:
                        if verbose:
                            print(f"    ✅ Asset gefunden: {asset_name} (ID: {asset.get('id', {}).get('id', 'N/A')})")
                        return asset
                
                # Prüfe ob es weitere Seiten gibt
                has_next = data.get('hasNext', False)
                page += 1
                
                if not has_next:
                    break
            else:
                print(f"    ⚠️  HTTP {response.status_code} beim Abrufen der Assets")
                if verbose:
                    print(f"    Response: {response.text}")
                break
        
        # Wenn nicht gefunden, versuche auch ohne textSearch (alle Assets durchsuchen)
        if verbose:
            print(f"    Asset '{asset_name}' nicht mit textSearch gefunden, durchsuche alle Assets...")
        
        page = 0
        has_next = True
        
        while has_next:
            params = {
                'page': page,
                'pageSize': page_size
            }
            
            response = requests.get(url, headers=headers, params=params, timeout=30)
            
            if response.status_code == 200:
                data = response.json()
                assets = data.get('data', [])
                
                for asset in assets:
                    asset_name_found = asset.get('name', '')
                    if asset_name_found == asset_name:
                        if verbose:
                            print(f"    ✅ Asset gefunden: {asset_name} (ID: {asset.get('id', {}).get('id', 'N/A')})")
                        return asset
                
                has_next = data.get('hasNext', False)
                page += 1
                
                if not has_next:
                    break
            else:
                break
        
        if verbose:
            print(f"    ❌ Asset '{asset_name}' nicht gefunden")
        
        return None
        
    except Exception as e:
        print(f"⚠️  Fehler beim Suchen nach Asset '{asset_name}': {e}")
        import traceback
        traceback.print_exc()
        return None

def get_asset_by_label_and_customer(label, customer_id, asset_profile_id=None, verbose=False):
    """Sucht ein Asset nach Label und Customer"""
    try:
        # Suche Assets des Customers
        url = f"{THINGSBOARD_BASE_URL}/api/customer/{customer_id}/assets"
        
        headers = {
            'X-Authorization': f'Bearer {token}',
            'Content-Type': 'application/json'
        }
        
        page = 0
        page_size = 100
        has_next = True
        
        while has_next:
            params = {
                'page': page,
                'pageSize': page_size
            }
            
            response = requests.get(url, headers=headers, params=params, timeout=30)
            
            if response.status_code == 200:
                data = response.json()
                assets = data.get('data', [])
                
                # Suche nach Label-Übereinstimmung (exakt)
                for asset in assets:
                    asset_label = asset.get('label', '')
                    if str(asset_label).strip() == str(label).strip():
                        # Optional: Prüfe auch Asset-Profil
                        if asset_profile_id:
                            asset_profile = asset.get('assetProfileId', {})
                            if isinstance(asset_profile, dict):
                                profile_id = asset_profile.get('id', '')
                                if profile_id == asset_profile_id:
                                    if verbose:
                                        print(f"    ✅ Asset gefunden nach Label '{label}' und Profil (ID: {asset.get('id', {}).get('id', 'N/A')})")
                                    return asset
                        # Wenn kein Profil-Match erforderlich oder Label passt, verwende es
                        if verbose:
                            print(f"    ✅ Asset gefunden nach Label '{label}' (ID: {asset.get('id', {}).get('id', 'N/A')})")
                        return asset
                
                # Prüfe ob es weitere Seiten gibt
                has_next = data.get('hasNext', False)
                page += 1
                
                if not has_next:
                    break
            else:
                if verbose:
                    print(f"    ⚠️  HTTP {response.status_code} beim Abrufen der Customer-Assets")
                break
        
        if verbose:
            print(f"    ❌ Kein Asset mit Label '{label}' gefunden")
        return None
        
    except Exception as e:
        if verbose:
            print(f"  ⚠️  Fehler beim Suchen nach Asset mit Label '{label}': {e}")
        return None

def get_all_customer_assets(customer_id):
    """Holt alle Assets eines Customers"""
    try:
        url = f"{THINGSBOARD_BASE_URL}/api/customer/{customer_id}/assets"
        headers = {
            'X-Authorization': f'Bearer {token}',
            'Content-Type': 'application/json'
        }
        
        all_assets = []
        page = 0
        page_size = 100
        
        while True:
            params = {
                'page': page,
                'pageSize': page_size
            }
            
            response = requests.get(url, headers=headers, params=params, timeout=30)
            
            if response.status_code == 200:
                data = response.json()
                assets = data.get('data', [])
                all_assets.extend(assets)
                
                has_next = data.get('hasNext', False)
                if not has_next:
                    break
                page += 1
            else:
                break
        
        return all_assets
        
    except Exception as e:
        print(f"  ⚠️  Fehler beim Abrufen der Customer-Assets: {e}")
        return []

def get_asset_profile_by_name(profile_name):
    """Sucht ein Asset-Profil nach Namen"""
    try:
        url = f"{THINGSBOARD_BASE_URL}/api/assetProfiles"
        headers = {
            'X-Authorization': f'Bearer {token}',
            'Content-Type': 'application/json'
        }
        
        params = {
            'pageSize': 100,
            'page': 0,
            'textSearch': profile_name
        }
        
        response = requests.get(url, headers=headers, params=params, timeout=30)
        
        if response.status_code == 200:
            data = response.json()
            profiles = data.get('data', [])
            
            for profile in profiles:
                if profile.get('name') == profile_name:
                    return profile
            
        return None
        
    except Exception as e:
        print(f"  ⚠️  Fehler beim Suchen nach Asset-Profil '{profile_name}': {e}")
        return None

def get_customer_info(customer_id):
    """Holt Informationen über einen Customer"""
    try:
        url = f"{THINGSBOARD_BASE_URL}/api/customer/{customer_id}"
        headers = {
            'X-Authorization': f'Bearer {token}',
            'Content-Type': 'application/json'
        }
        
        response = requests.get(url, headers=headers, timeout=30)
        
        if response.status_code == 200:
            return response.json()
        else:
            print(f"  ⚠️  Customer mit ID '{customer_id}' nicht gefunden: {response.status_code}")
            return None
            
    except Exception as e:
        print(f"  ⚠️  Fehler beim Abrufen der Customer-Info: {e}")
        return None

def create_asset(name, label, asset_type='ASSET', asset_profile_id=None, customer_id=None, dry_run=False, max_retries=3, retry_delay=2):
    """Erstellt ein Asset in Thingsboard mit Retry-Logik"""
    import time
    
    try:
        if dry_run:
            profile_info = f" (Profil-ID: {asset_profile_id})" if asset_profile_id else ""
            customer_info = f" (Customer-ID: {customer_id})" if customer_id else ""
            print(f"  [DRY-RUN] Würde Asset erstellen: {name} (Label: {label}){profile_info}{customer_info}")
            return {'id': {'id': f'DRY_RUN_{name}'}, 'name': name, 'label': label}
        
        url = f"{THINGSBOARD_BASE_URL}/api/asset"
        
        asset_data = {
            "name": name,
            "type": asset_type,
            "label": label
        }
        
        # Füge Asset-Profil hinzu, falls angegeben
        if asset_profile_id:
            asset_data["assetProfileId"] = {
                "id": asset_profile_id,
                "entityType": "ASSET_PROFILE"
            }
        
        # Füge Customer hinzu, falls angegeben
        if customer_id:
            asset_data["customerId"] = {
                "id": customer_id,
                "entityType": "CUSTOMER"
            }
        
        headers = {
            'X-Authorization': f'Bearer {token}',
            'Content-Type': 'application/json'
        }
        
        # Retry-Logik für Rate Limiting und temporäre Fehler
        for attempt in range(max_retries):
            try:
                response = requests.post(url, json=asset_data, headers=headers, timeout=30)
                
                if response.status_code == 200:
                    asset = response.json()
                    profile_info = f" (Profil-ID: {asset_profile_id})" if asset_profile_id else ""
                    customer_info = f" (Customer-ID: {customer_id})" if customer_id else ""
                    print(f"  ✅ Asset erstellt: {name} (ID: {asset.get('id', {}).get('id', 'N/A')}){profile_info}{customer_info}")
                    return asset
                elif response.status_code == 429:  # Rate Limiting
                    if attempt < max_retries - 1:
                        wait_time = retry_delay * (attempt + 1)
                        print(f"  ⚠️  Rate Limit erreicht (Versuch {attempt + 1}/{max_retries}), warte {wait_time}s...")
                        time.sleep(wait_time)
                        continue
                    else:
                        print(f"  ❌ Rate Limit nach {max_retries} Versuchen: {response.status_code}")
                        print(f"     Response: {response.text}")
                        return None
                elif response.status_code == 400:  # Bad Request - möglicherweise Duplikat
                    error_text = response.text.lower()
                    if 'already exists' in error_text or 'duplicate' in error_text or 'name' in error_text:
                        # Prüfe ob Asset bereits existiert
                        existing_asset = get_asset_by_name(name)
                        if existing_asset:
                            print(f"  ℹ️  Asset '{name}' existiert bereits, verwende existierendes Asset")
                            return existing_asset
                        else:
                            print(f"  ⚠️  Asset-Name möglicherweise bereits vergeben: {response.status_code}")
                            print(f"     Response: {response.text}")
                            return None
                    else:
                        print(f"  ❌ Fehler beim Erstellen von Asset '{name}': {response.status_code}")
                        print(f"     Response: {response.text}")
                        return None
                elif response.status_code >= 500:  # Server-Fehler - retry
                    if attempt < max_retries - 1:
                        wait_time = retry_delay * (attempt + 1)
                        print(f"  ⚠️  Server-Fehler {response.status_code} (Versuch {attempt + 1}/{max_retries}), warte {wait_time}s...")
                        time.sleep(wait_time)
                        continue
                    else:
                        print(f"  ❌ Server-Fehler nach {max_retries} Versuchen: {response.status_code}")
                        print(f"     Response: {response.text}")
                        return None
                else:
                    print(f"  ❌ Fehler beim Erstellen von Asset '{name}': {response.status_code}")
                    print(f"     Response: {response.text}")
                    return None
                    
            except requests.exceptions.Timeout:
                if attempt < max_retries - 1:
                    wait_time = retry_delay * (attempt + 1)
                    print(f"  ⚠️  Timeout (Versuch {attempt + 1}/{max_retries}), warte {wait_time}s...")
                    time.sleep(wait_time)
                    continue
                else:
                    print(f"  ❌ Timeout nach {max_retries} Versuchen")
                    return None
            except requests.exceptions.ConnectionError:
                if attempt < max_retries - 1:
                    wait_time = retry_delay * (attempt + 1)
                    print(f"  ⚠️  Verbindungsfehler (Versuch {attempt + 1}/{max_retries}), warte {wait_time}s...")
                    time.sleep(wait_time)
                    continue
                else:
                    print(f"  ❌ Verbindungsfehler nach {max_retries} Versuchen")
                    return None
            
        return None
            
    except Exception as e:
        print(f"  ❌ Fehler beim Erstellen von Asset '{name}': {e}")
        import traceback
        traceback.print_exc()
        return None

def create_relation(from_asset_id, to_asset_id, relation_type='Contains', dry_run=False):
    """Erstellt eine Relation zwischen zwei Assets"""
    try:
        if dry_run:
            print(f"    [DRY-RUN] Würde Relation erstellen: {from_asset_id} → {to_asset_id} ({relation_type})")
            return True
        
        url = f"{THINGSBOARD_BASE_URL}/api/relation"
        
        relation_data = {
            "from": {
                "entityType": "ASSET",
                "id": from_asset_id
            },
            "to": {
                "entityType": "ASSET",
                "id": to_asset_id
            },
            "type": relation_type,
            "typeGroup": "COMMON"
        }
        
        headers = {
            'X-Authorization': f'Bearer {token}',
            'Content-Type': 'application/json'
        }
        
        response = requests.post(url, json=relation_data, headers=headers, timeout=30)
        
        if response.status_code == 200:
            print(f"    ✅ Relation erstellt: {from_asset_id} → {to_asset_id}")
            return True
        else:
            # Prüfe ob Relation bereits existiert
            if response.status_code == 400:
                error_text = response.text.lower()
                if 'already exists' in error_text or 'duplicate' in error_text:
                    print(f"    ℹ️  Relation existiert bereits: {from_asset_id} → {to_asset_id}")
                    return True
            
            print(f"    ⚠️  Fehler beim Erstellen der Relation: {response.status_code}")
            print(f"       Response: {response.text}")
            return False
            
    except Exception as e:
        print(f"    ❌ Fehler beim Erstellen der Relation: {e}")
        return False

def load_excel_file(excel_file_path):
    """Lädt die Excel-Datei und sortiert nach Level 1-6"""
    try:
        print(f"📁 Lade Excel-Datei: {excel_file_path}")
        
        if not os.path.exists(excel_file_path):
            print(f"❌ Datei nicht gefunden: {excel_file_path}")
            return None
        
        # Lade Excel-Datei: 
        # - Zeile 1 = Header (wird verwendet)
        # - Zeile 2 = wird übersprungen
        # - Zeile 3+ = Daten
        df = pd.read_excel(excel_file_path, header=0, skiprows=[1])
        print(f"✅ Excel-Datei geladen: {len(df)} Zeilen (ab Zeile 3)")
        print(f"📊 Spalten: {list(df.columns)}")
        
        # Normalisiere Spaltennamen (entferne führende/nachfolgende Leerzeichen)
        df.columns = df.columns.str.strip()
        
        # Prüfe ob benötigte Spalten vorhanden sind (mit flexibler Suche)
        required_columns = ['Level 1', 'Level 2', 'Level 3', 'Level 4', 'Level 5', 'Level 6']
        missing_columns = []
        column_mapping = {}
        
        for req_col in required_columns:
            # Suche nach exakter Übereinstimmung oder mit Leerzeichen
            found = False
            for actual_col in df.columns:
                if actual_col.strip() == req_col:
                    column_mapping[req_col] = actual_col
                    found = True
                    break
            if not found:
                missing_columns.append(req_col)
        
        if missing_columns:
            print(f"❌ Fehlende Spalten: {missing_columns}")
            print(f"   Verfügbare Spalten: {list(df.columns)}")
            return None
        
        # Normalisiere Spaltennamen für weitere Verarbeitung
        for req_col, actual_col in column_mapping.items():
            if req_col != actual_col:
                df.rename(columns={actual_col: req_col}, inplace=True)
        
        # Sortiere nach Level 1-6 (von oben nach unten in der Hierarchie)
        sort_columns = ['Level 1', 'Level 2', 'Level 3', 'Level 4', 'Level 5', 'Level 6']
        available_sort_columns = [col for col in sort_columns if col in df.columns]
        
        if available_sort_columns:
            print(f"🔄 Sortiere nach: {', '.join(available_sort_columns)}")
            df = df.sort_values(by=available_sort_columns, na_position='last')
            print(f"✅ Daten sortiert")
        else:
            print(f"⚠️  Sortierspalten nicht gefunden, verwende Standardreihenfolge")
        
        return df
        
    except Exception as e:
        print(f"❌ Fehler beim Laden der Excel-Datei: {e}")
        import traceback
        traceback.print_exc()
        return None

def is_valid_value(value):
    """Prüft ob ein Wert gültig ist (nicht None, NaN oder leer)"""
    import pandas as pd
    if value is None:
        return False
    if pd.isna(value):
        return False
    value_str = str(value).strip()
    if value_str == '' or value_str.lower() == 'nan':
        return False
    return True

def process_assets(df, dry_run=False):
    """Verarbeitet die Excel-Daten und erstellt Assets mit Relations"""
    
    # Prüfe ob Root Asset existiert
    root_asset_name = "JHRZO_0001"
    print(f"\n🔍 Suche nach Root Asset: {root_asset_name}")
    
    if dry_run:
        root_asset_id = "ROOT_ASSET_ID"
        print(f"  [DRY-RUN] Verwende Root Asset: {root_asset_name}")
    else:
        root_asset = get_asset_by_name(root_asset_name, verbose=True)
        
        if not root_asset:
            print(f"❌ Root Asset '{root_asset_name}' nicht gefunden!")
            print(f"   Bitte prüfen Sie:")
            print(f"   1. Ob das Asset in Thingsboard existiert")
            print(f"   2. Ob der Name exakt '{root_asset_name}' ist (ohne Leerzeichen)")
            print(f"   3. Ob Sie die richtigen Berechtigungen haben")
            return None
        
        root_asset_id = root_asset.get('id', {}).get('id')
        print(f"✅ Root Asset gefunden: {root_asset_name} (ID: {root_asset_id})")
    
    # Customer-ID - Johanniter Customer
    customer_id = "8d594810-d431-11f0-a316-ddf16a8a4c6c"
    print(f"✅ Customer-ID gesetzt: {customer_id}")
    
    # Excel-Struktur:
    # Spalte A = Sensor DevEUI (wird ignoriert)
    # Spalte B = Level 6 (Root, bereits importiert als JHRZO_0001)
    # Spalte C = Level 5 (City)
    # Spalte D = Level 4 (Building)
    # Spalte E = Level 3 (Floor)
    # Spalte F = Level 2 (Area)
    # Spalte G = Level 1 (Room)
    # Spalte H = Ebene 6 (Asset-Name Level 6, wird geschrieben)
    # Spalte I = Ebene 5 (Asset-Name Level 5, wird geschrieben)
    # Spalte J = Ebene 4 (Asset-Name Level 4, wird geschrieben)
    # Spalte K = Ebene 3 (Asset-Name Level 3, wird geschrieben)
    # Spalte L = Ebene 2 (Asset-Name Level 2, wird geschrieben)
    # Spalte M = Ebene 1 (Asset-Name Level 1, wird geschrieben)
    
    # Hole Asset-Profile IDs
    print(f"\n🔍 Suche nach Asset-Profilen...")
    city_profile = None
    building_profile = None
    floor_profile = None
    area_profile = None
    room_profile = None
    
    if not dry_run:
        city_profile = get_asset_profile_by_name("City")
        building_profile = get_asset_profile_by_name("Building")
        floor_profile = get_asset_profile_by_name("Floor")
        area_profile = get_asset_profile_by_name("Area")
        room_profile = get_asset_profile_by_name("Room")
        
        if city_profile:
            print(f"  ✅ Asset-Profil 'City' gefunden (ID: {city_profile.get('id', {}).get('id', 'N/A')})")
        else:
            print(f"  ⚠️  Asset-Profil 'City' nicht gefunden")
        
        if building_profile:
            print(f"  ✅ Asset-Profil 'Building' gefunden (ID: {building_profile.get('id', {}).get('id', 'N/A')})")
        else:
            print(f"  ⚠️  Asset-Profil 'Building' nicht gefunden")
        
        if floor_profile:
            print(f"  ✅ Asset-Profil 'Floor' gefunden (ID: {floor_profile.get('id', {}).get('id', 'N/A')})")
        else:
            print(f"  ⚠️  Asset-Profil 'Floor' nicht gefunden")
        
        if area_profile:
            print(f"  ✅ Asset-Profil 'Area' gefunden (ID: {area_profile.get('id', {}).get('id', 'N/A')})")
        else:
            print(f"  ⚠️  Asset-Profil 'Area' nicht gefunden")
        
        if room_profile:
            print(f"  ✅ Asset-Profil 'Room' gefunden (ID: {room_profile.get('id', {}).get('id', 'N/A')})")
        else:
            print(f"  ⚠️  Asset-Profil 'Room' nicht gefunden")
    
    city_profile_id = city_profile.get('id', {}).get('id') if city_profile else None
    building_profile_id = building_profile.get('id', {}).get('id') if building_profile else None
    floor_profile_id = floor_profile.get('id', {}).get('id') if floor_profile else None
    area_profile_id = area_profile.get('id', {}).get('id') if area_profile else None
    room_profile_id = room_profile.get('id', {}).get('id') if room_profile else None
    
    # Erstelle Hierarchie-Struktur
    # Level-Struktur: Level 6 (root, JHRZO_0001) → Level 5 (City) → Level 4 (Building) → Level 3 (Floor) → Level 2 (Area) → Level 1 (Room)
    
    # Bestimme Start-Counter: Suche nach höchster existierender JHRZO_XXXX Nummer
    asset_counter = 2  # Default: Start bei JHRZO_0002
    if not dry_run:
        print(f"\n🔍 Bestimme Start-Counter aus existierenden Assets...")
        try:
            # Suche nach allen JHRZO_ Assets
            url = f"{THINGSBOARD_BASE_URL}/api/tenant/assets"
            headers = {
                'X-Authorization': f'Bearer {token}',
                'Content-Type': 'application/json'
            }
            
            page = 0
            page_size = 100
            has_next = True
            max_number = 1
            
            while has_next:
                params = {
                    'page': page,
                    'pageSize': page_size,
                    'textSearch': 'JHRZO_'
                }
                
                response = requests.get(url, headers=headers, params=params, timeout=30)
                
                if response.status_code == 200:
                    data = response.json()
                    assets = data.get('data', [])
                    
                    for asset in assets:
                        asset_name = asset.get('name', '')
                        if asset_name.startswith('JHRZO_'):
                            try:
                                # Extrahiere Nummer aus JHRZO_XXXX
                                number_str = asset_name.replace('JHRZO_', '').strip()
                                number = int(number_str)
                                if number > max_number:
                                    max_number = number
                            except (ValueError, AttributeError):
                                continue
                    
                    has_next = data.get('hasNext', False)
                    page += 1
                    
                    if not has_next:
                        break
                else:
                    break
            
            if max_number >= 1:
                asset_counter = max_number + 1
                print(f"  ✅ Höchste gefundene Asset-Nummer: JHRZO_{max_number:05d}")
                print(f"  ✅ Starte bei Asset-Nummer: JHRZO_{asset_counter:05d}")
            else:
                print(f"  ℹ️  Keine JHRZO_ Assets gefunden, starte bei JHRZO_00002")
        except Exception as e:
            print(f"  ⚠️  Fehler beim Bestimmen des Start-Counters: {e}")
            print(f"  ℹ️  Verwende Standard-Start: JHRZO_00002")
    
    # Initialisiere Spalten H, I, J, K, L, M falls nicht vorhanden (für Asset-Namen)
    # Spalte H = Ebene 1 (Level 1 Asset-Name)
    # Spalte I = Ebene 2 (Level 2 Asset-Name)
    # Spalte J = Ebene 3 (Level 3 Asset-Name)
    # Spalte K = Ebene 4 (Level 4 Asset-Name)
    # Spalte L = Ebene 5 (Level 5 Asset-Name)
    # Spalte M = Ebene 6 (Level 6 Asset-Name)
    output_columns = ['Ebene 1', 'Ebene 2', 'Ebene 3', 'Ebene 4', 'Ebene 5', 'Ebene 6']
    for col_name in output_columns:
        if col_name not in df.columns:
            df[col_name] = ''
    
    # Lade existierende Assets (nur Customer-Assets oder Assets mit JHRZO im Namen)
    existing_assets_by_label = {}
    if not dry_run:
        print(f"\n🔍 Lade existierende Assets (Customer-Assets oder Assets mit JHRZO im Namen)...")
        headers = {
            'X-Authorization': f'Bearer {token}',
            'Content-Type': 'application/json'
        }
        
        # 1. Lade Customer-Assets
        print(f"  📋 Lade Assets des Customers {customer_id}...")
        customer_url = f"{THINGSBOARD_BASE_URL}/api/customer/{customer_id}/assets"
        page = 0
        page_size = 100
        has_next = True
        customer_assets_count = 0
        
        while has_next:
            params = {
                'page': page,
                'pageSize': page_size
            }
            
            response = requests.get(customer_url, headers=headers, params=params, timeout=30)
            
            if response.status_code == 200:
                data = response.json()
                assets = data.get('data', [])
                
                for asset in assets:
                    label = asset.get('label', '')
                    if label:
                        existing_assets_by_label[label] = asset
                        customer_assets_count += 1
                
                has_next = data.get('hasNext', False)
                page += 1
                
                if not has_next:
                    break
            else:
                print(f"  ⚠️  Fehler beim Laden der Customer-Assets: {response.status_code}")
                break
        
        print(f"  ✅ {customer_assets_count} Customer-Assets geladen")
        
        # 2. Lade Assets mit JHRZO im Namen (falls nicht bereits geladen)
        print(f"  📋 Lade Assets mit 'JHRZO' im Namen...")
        tenant_url = f"{THINGSBOARD_BASE_URL}/api/tenant/assets"
        page = 0
        has_next = True
        jhrzo_assets_count = 0
        
        while has_next:
            params = {
                'page': page,
                'pageSize': page_size,
                'textSearch': 'JHRZO'
            }
            
            response = requests.get(tenant_url, headers=headers, params=params, timeout=30)
            
            if response.status_code == 200:
                data = response.json()
                assets = data.get('data', [])
                
                for asset in assets:
                    asset_name = asset.get('name', '')
                    # Prüfe ob Asset mit JHRZO beginnt
                    if asset_name and asset_name.startswith('JHRZO'):
                        label = asset.get('label', '')
                        if label:
                            # Überschreibe nur wenn noch nicht vorhanden (Customer-Assets haben Priorität)
                            if label not in existing_assets_by_label:
                                existing_assets_by_label[label] = asset
                                jhrzo_assets_count += 1
                
                has_next = data.get('hasNext', False)
                page += 1
                
                if not has_next:
                    break
            else:
                print(f"  ⚠️  Fehler beim Laden der JHRZO-Assets: {response.status_code}")
                break
        
        print(f"  ✅ {jhrzo_assets_count} zusätzliche JHRZO-Assets geladen")
        print(f"  ✅ Gesamt: {len(existing_assets_by_label)} Assets mit Labels geladen")
    
    def process_level(level_num, level_column, parent_assets_dict, parent_key_func, asset_profile_id, parent_level_name):
        """Verarbeitet ein Level der Hierarchie"""
        nonlocal asset_counter  # Zugriff auf asset_counter aus der äußeren Funktion
        
        level_name = f"Level {level_num}"
        print(f"\n📋 Verarbeite {level_name} (Spalte {level_column})...")
        print(f"   Parent-Level: {parent_level_name}")
        print(f"   Asset-Profil-ID: {asset_profile_id}")
        print(f"   Parent-Assets im Dictionary: {len(parent_assets_dict)}")
        
        level_assets = {}  # {key: asset_name}
        created_count = 0
        found_count = 0
        
        # Sammle eindeutige Werte für dieses Level
        unique_values = set()
        total_rows = 0
        rows_with_value = 0
        rows_with_parent = 0
        
        for _, row in df.iterrows():
            total_rows += 1
            level_value = str(row.get(level_column, '')).strip()
            if level_value and level_value != 'nan':
                rows_with_value += 1
                # Erstelle Key basierend auf Parent-Levels
                if parent_key_func:
                    parent_key = parent_key_func(row)
                    if parent_key:
                        rows_with_parent += 1
                        if parent_key in parent_assets_dict:
                            unique_values.add((parent_key, level_value))
                        else:
                            print(f"   ⚠️  Parent-Key '{parent_key}' nicht in parent_assets_dict gefunden (Level-Wert: '{level_value}')")
                    else:
                        print(f"   ⚠️  Kein Parent-Key für Zeile mit Level-Wert '{level_value}'")
                else:
                    # Level 2 hat kein Parent (außer Root)
                    unique_values.add((None, level_value))
        
        print(f"   📊 Statistik:")
        print(f"      Gesamt Zeilen: {total_rows}")
        print(f"      Zeilen mit {level_column}-Wert: {rows_with_value}")
        if parent_key_func:
            print(f"      Zeilen mit Parent-Key: {rows_with_parent}")
        print(f"      Eindeutige Werte gefunden: {len(unique_values)}")
        
        if len(unique_values) == 0:
            print(f"   ⚠️  KEINE eindeutigen Werte gefunden für {level_name}!")
            print(f"      Mögliche Ursachen:")
            print(f"      - Spalte '{level_column}' existiert nicht oder ist leer")
            print(f"      - Parent-Assets fehlen für alle Zeilen")
            return {}
        
        print(f"   📝 Eindeutige Werte:")
        for idx, (uk, lv) in enumerate(sorted(unique_values), 1):
            print(f"      {idx}. Parent-Key: {uk}, Level-Wert: '{lv}'")
        
        processed_count = 0
        for unique_key, level_value in unique_values:
            processed_count += 1
            parent_key = unique_key
            asset_label = level_value
            
            print(f"\n   🔍 Verarbeite Wert {processed_count}/{len(unique_values)}: '{asset_label}'")
            if parent_key:
                print(f"      Parent-Key: {parent_key}")
            
            # Prüfe zuerst ob in Output-Spalte bereits ein Asset-Name steht
            existing_asset_name = None
            # Output-Spalten: Ebene 6 (H), Ebene 5 (I), Ebene 4 (J), Ebene 3 (K), Ebene 2 (L), Ebene 1 (M)
            # Mapping: Level 6 → Ebene 6 (H), Level 5 → Ebene 5 (I), Level 4 → Ebene 4 (J), Level 3 → Ebene 3 (K), Level 2 → Ebene 2 (L), Level 1 → Ebene 1 (M)
            output_column_names = ['Ebene 6', 'Ebene 5', 'Ebene 4', 'Ebene 3', 'Ebene 2', 'Ebene 1']
            # Level 6 → Index 0, Level 5 → Index 1, Level 4 → Index 2, Level 3 → Index 3, Level 2 → Index 4, Level 1 → Index 5
            output_col_name = output_column_names[6 - level_num] if 1 <= level_num <= 6 else None
            
            print(f"      Prüfe Output-Spalte '{output_col_name}'...")
            matching_rows = 0
            for idx, row in df.iterrows():
                row_level_value = str(row.get(level_column, '')).strip()
                if row_level_value == level_value:
                    matching_rows += 1
                    # Prüfe Parent-Match falls vorhanden
                    if parent_key_func:
                        row_parent_key = parent_key_func(row)
                        if row_parent_key != parent_key:
                            continue
                    
                    # Prüfe Output-Spalte
                    if output_col_name and output_col_name in df.columns:
                        output_value = row.get(output_col_name, '')
                        # Prüfe ob Wert wirklich vorhanden ist (nicht NaN, None oder leer)
                        if is_valid_value(output_value):
                            existing_asset_name = str(output_value).strip()
                            print(f"      ✅ Asset-Name in Excel gefunden: {existing_asset_name}")
                            break
            
            if matching_rows == 0:
                print(f"      ⚠️  Keine passenden Zeilen gefunden für '{asset_label}'")
            
            if not existing_asset_name:
                print(f"      ℹ️  Kein Asset-Name in Excel-Spalte gefunden, suche nach existierendem Asset oder erstelle neues")
            
            existing_asset = None
            asset_name = None
            
            if existing_asset_name:
                # Prüfe ob Asset mit diesem Namen existiert
                print(f"      🔍 Suche nach Asset mit Namen '{existing_asset_name}'...")
                if not dry_run:
                    existing_asset = get_asset_by_name(existing_asset_name)
                if existing_asset:
                    # Prüfe ob das Label des Assets mit dem erwarteten Label übereinstimmt
                    existing_label = existing_asset.get('label', '').strip()
                    if existing_label == asset_label:
                        # Label passt, verwende existierendes Asset
                        asset_name = existing_asset_name
                        print(f"      ✅ Asset aus Excel gefunden: {asset_name} (Label: {asset_label})")
                        found_count += 1
                    else:
                        # Label passt nicht - das Asset gehört nicht zu diesem Label
                        print(f"      ⚠️  Asset '{existing_asset_name}' hat Label '{existing_label}', erwartet '{asset_label}'")
                        print(f"      ℹ️  Erstelle neues Asset für Label '{asset_label}'")
                        existing_asset_name = None  # Setze zurück, damit neues Asset erstellt wird
                        existing_asset = None
                else:
                    print(f"      ⚠️  Asset mit Namen '{existing_asset_name}' nicht gefunden")
                    existing_asset_name = None  # Setze zurück, damit neues Asset erstellt wird
            else:
                # Suche nach Asset mit gleichem Label
                print(f"      🔍 Suche nach Asset mit Label '{asset_label}'...")
                if not dry_run:
                    existing_asset = existing_assets_by_label.get(asset_label)
                    if existing_asset:
                        print(f"      ℹ️  Asset mit Label gefunden, prüfe Profil...")
                        # Prüfe ob Profil passt
                        if asset_profile_id:
                            existing_profile_id = existing_asset.get('assetProfileId', {})
                            if isinstance(existing_profile_id, dict):
                                existing_profile_id = existing_profile_id.get('id', '')
                            print(f"         Existierendes Profil-ID: {existing_profile_id}")
                            print(f"         Erwartetes Profil-ID: {asset_profile_id}")
                            if existing_profile_id != asset_profile_id:
                                existing_asset = None  # Profil passt nicht
                                print(f"         ⚠️  Profil passt nicht, ignoriere Asset")
                            else:
                                print(f"         ✅ Profil passt")
                
                if existing_asset:
                    asset_name = existing_asset.get('name', '')
                    print(f"      ✅ Existierendes Asset gefunden: {asset_name} (Label: {asset_label})")
                    found_count += 1
                else:
                    # Erstelle neues Asset
                    print(f"      🔨 Erstelle neues Asset...")
                    asset_name = f"JHRZO_{asset_counter:05d}"
                    print(f"         Asset-Name: {asset_name}")
                    print(f"         Label: {asset_label}")
                    print(f"         Profil-ID: {asset_profile_id}")
                    print(f"         Customer-ID: {customer_id}")
                    
                    asset = create_asset(asset_name, asset_label, asset_profile_id=asset_profile_id, customer_id=customer_id, dry_run=dry_run)
                    if asset:
                        asset_id = asset.get('id', {}).get('id') if isinstance(asset.get('id'), dict) else asset.get('id', f'DRY_RUN_{asset_name}')
                        print(f"         Asset-ID: {asset_id}")
                        
                        # Relation zu Parent Asset
                        parent_asset_name = None
                        if level_num == 5:
                            # Level 5 hat Root Asset (Level 6) als Parent
                            print(f"         Erstelle Relation zu Root Asset Level 6 ({root_asset_id})...")
                            create_relation(root_asset_id, asset_id, dry_run=dry_run)
                        elif parent_key and parent_assets_dict:
                            # Für Level 3+: parent_key ist der Parent-Key (z.B. level2_val für Level 3)
                            if parent_key in parent_assets_dict:
                                parent_asset_name = parent_assets_dict[parent_key]
                                print(f"         Parent-Asset-Name: {parent_asset_name}")
                                parent_asset_obj = get_asset_by_name(parent_asset_name) if not dry_run else None
                                if parent_asset_obj:
                                    parent_asset_id = parent_asset_obj.get('id', {}).get('id')
                                    print(f"         Erstelle Relation zu Parent Asset ({parent_asset_id})...")
                                    create_relation(parent_asset_id, asset_id, dry_run=dry_run)
                                else:
                                    print(f"         ⚠️  Parent Asset '{parent_asset_name}' nicht gefunden")
                            else:
                                print(f"         ⚠️  Parent-Key '{parent_key}' nicht in parent_assets_dict gefunden")
                        
                        asset_counter += 1
                        created_count += 1
                        print(f"      ✅ Asset erfolgreich erstellt: {asset_name}")
                    else:
                        # Prüfe ob Asset vielleicht doch existiert
                        if not dry_run:
                            print(f"      ⚠️  Asset-Erstellung fehlgeschlagen, prüfe ob Asset bereits existiert...")
                            existing_check = get_asset_by_name(asset_name)
                            if existing_check:
                                print(f"      ℹ️  Asset '{asset_name}' existiert bereits")
                                asset_counter += 1
                                found_count += 1
                            else:
                                print(f"      ❌ FEHLER: Konnte Asset '{asset_name}' nicht erstellen!")
                                asset_counter += 1
                                continue
                        else:
                            asset_counter += 1
                            continue
                    
                    # Kleine Pause zwischen Requests
                    if not dry_run and created_count % 10 == 0:
                        import time
                        print(f"      ⏸️  Pause nach {created_count} erstellten Assets...")
                        time.sleep(0.5)
            
            if asset_name:
                # Speichere Asset-Name mit dem richtigen Key
                if level_num == 5:
                    # Level 5: Key ist nur der Wert (hat Root als Parent)
                    level_assets[level_value] = asset_name
                    print(f"      💾 Gespeichert: Key='{level_value}' → Asset='{asset_name}'")
                else:
                    # Level 3+: Key ist das vollständige Tuple (parent_key, level_value)
                    # parent_key ist bereits der Parent-Key (z.B. level2_val für Level 3)
                    # Wir müssen den vollständigen Key erstellen für das nächste Level
                    if parent_key:
                        if isinstance(parent_key, tuple):
                            full_key = parent_key + (level_value,)
                        else:
                            full_key = (parent_key, level_value)
                        level_assets[full_key] = asset_name
                        print(f"      💾 Gespeichert: Key={full_key} → Asset='{asset_name}'")
                    else:
                        level_assets[level_value] = asset_name
                        print(f"      💾 Gespeichert: Key='{level_value}' → Asset='{asset_name}'")
            else:
                print(f"      ⚠️  Kein Asset-Name für '{asset_label}' - wird übersprungen")
        
        print(f"\n   📊 Zusammenfassung {level_name}:")
        print(f"      Gefunden: {found_count}")
        print(f"      Erstellt: {created_count}")
        print(f"      Gesamt im Dictionary: {len(level_assets)}")
        if len(level_assets) > 0:
            print(f"      Asset-Namen: {list(level_assets.values())[:5]}{'...' if len(level_assets) > 5 else ''}")
        
        return level_assets
    
    # Verarbeite Level 5 (City) - Parent: Level 6 (Root Asset JHRZO_0001)
    level5_assets = process_level(5, 'Level 5', {}, None, city_profile_id, 'Level 6')  # Ebene 5 (Spalte I)
    
    # Verarbeite Level 4 (Building) - Parent: Level 5
    def get_level5_parent_key(row):
        level5_val = str(row.get('Level 5', '')).strip()
        return level5_val if level5_val and level5_val != 'nan' else None
    
    level4_assets = process_level(4, 'Level 4', level5_assets, get_level5_parent_key, building_profile_id, 'Level 5')  # Ebene 4 (Spalte J)
    
    # Verarbeite Level 3 (Floor) - Parent: Level 4
    def get_level4_parent_key(row):
        level5_val = str(row.get('Level 5', '')).strip()
        level4_val = str(row.get('Level 4', '')).strip()
        if level5_val and level5_val != 'nan' and level4_val and level4_val != 'nan':
            return (level5_val, level4_val)
        return None
    
    level3_assets = process_level(3, 'Level 3', level4_assets, get_level4_parent_key, floor_profile_id, 'Level 4')  # Ebene 3 (Spalte K)
    
    # Verarbeite Level 2 (Area) - Parent: Level 3
    def get_level3_parent_key(row):
        level5_val = str(row.get('Level 5', '')).strip()
        level4_val = str(row.get('Level 4', '')).strip()
        level3_val = str(row.get('Level 3', '')).strip()
        if level5_val and level5_val != 'nan' and level4_val and level4_val != 'nan' and level3_val and level3_val != 'nan':
            return (level5_val, level4_val, level3_val)
        return None
    
    level2_assets = process_level(2, 'Level 2', level3_assets, get_level3_parent_key, area_profile_id, 'Level 3')  # Ebene 2 (Spalte L)
    
    # Verarbeite Level 1 (Room) - Parent: Level 2
    def get_level2_parent_key(row):
        level5_val = str(row.get('Level 5', '')).strip()
        level4_val = str(row.get('Level 4', '')).strip()
        level3_val = str(row.get('Level 3', '')).strip()
        level2_val = str(row.get('Level 2', '')).strip()
        if all(v and v != 'nan' for v in [level5_val, level4_val, level3_val, level2_val]):
            return (level5_val, level4_val, level3_val, level2_val)
        return None
    
    level1_assets = process_level(1, 'Level 1', level2_assets, get_level2_parent_key, room_profile_id, 'Level 2')  # Ebene 1 (Spalte M)
    
    # Schreibe Asset-Namen in Spalten Ebene 6-1 (H, I, J, K, L, M)
    print(f"\n📝 Schreibe Asset-Namen in Spalten Ebene 6-1 (H, I, J, K, L, M)...")
    for idx, row in df.iterrows():
        level1_val = str(row.get('Level 1', '')).strip()
        level2_val = str(row.get('Level 2', '')).strip()
        level3_val = str(row.get('Level 3', '')).strip()
        level4_val = str(row.get('Level 4', '')).strip()
        level5_val = str(row.get('Level 5', '')).strip()
        level6_val = str(row.get('Level 6', '')).strip()
        
        # Ebene 6 (Spalte H): Asset-Name für Level 6 (Root - bereits bekannt)
        if level6_val and level6_val != 'nan':
            df.at[idx, 'Ebene 6'] = root_asset_name
        
        # Ebene 5 (Spalte I): Asset-Name für Level 5 (City)
        if level5_val and level5_val != 'nan':
            if level5_val in level5_assets:
                df.at[idx, 'Ebene 5'] = level5_assets[level5_val]
        
        # Ebene 4 (Spalte J): Asset-Name für Level 4 (Building)
        if level5_val and level5_val != 'nan' and level4_val and level4_val != 'nan':
            key = (level5_val, level4_val)
            if key in level4_assets:
                df.at[idx, 'Ebene 4'] = level4_assets[key]
        
        # Ebene 3 (Spalte K): Asset-Name für Level 3 (Floor)
        if (level5_val and level5_val != 'nan' and 
            level4_val and level4_val != 'nan' and 
            level3_val and level3_val != 'nan'):
            key = (level5_val, level4_val, level3_val)
            if key in level3_assets:
                df.at[idx, 'Ebene 3'] = level3_assets[key]
        
        # Ebene 2 (Spalte L): Asset-Name für Level 2 (Area)
        if (level5_val and level5_val != 'nan' and 
            level4_val and level4_val != 'nan' and 
            level3_val and level3_val != 'nan' and
            level2_val and level2_val != 'nan'):
            key = (level5_val, level4_val, level3_val, level2_val)
            if key in level2_assets:
                df.at[idx, 'Ebene 2'] = level2_assets[key]
        
        # Ebene 1 (Spalte M): Asset-Name für Level 1 (Room)
        if (level5_val and level5_val != 'nan' and 
            level4_val and level4_val != 'nan' and 
            level3_val and level3_val != 'nan' and
            level2_val and level2_val != 'nan' and
            level1_val and level1_val != 'nan'):
            key = (level5_val, level4_val, level3_val, level2_val, level1_val)
            if key in level1_assets:
                df.at[idx, 'Ebene 1'] = level1_assets[key]
    
    print(f"✅ Asset-Namen in Spalten Ebene 6-1 geschrieben")
    
    print(f"\n✅ Verarbeitung abgeschlossen!")
    print(f"   Level 6 (Root): 1 (JHRZO_0001)")
    print(f"   Level 5 (City): {len(level5_assets)}")
    print(f"   Level 4 (Building): {len(level4_assets)}")
    print(f"   Level 3 (Floor): {len(level3_assets)}")
    print(f"   Level 2 (Area): {len(level2_assets)}")
    print(f"   Level 1 (Room): {len(level1_assets)}")
    print(f"   Nächste Asset-Nummer: JHRZO_{asset_counter:05d}")
    
    return df

def main():
    """Hauptfunktion"""
    import argparse
    
    parser = argparse.ArgumentParser(
        description='Importiert Assets in Thingsboard aus Excel-Datei',
        formatter_class=argparse.RawDescriptionHelpFormatter
    )
    
    parser.add_argument(
        '--excel-file',
        default='data/Struktur_johanniter.xlsx',
        help='Excel-Datei (Standard: data/Struktur_johanniter.xlsx)'
    )
    
    parser.add_argument(
        '--dry-run',
        action='store_true',
        help='Dry-Run Modus: Zeigt was erstellt würde, ohne tatsächlich zu erstellen'
    )
    
    parser.add_argument(
        '--verbose',
        action='store_true',
        help='Ausführliche Ausgabe'
    )
    
    parser.add_argument(
        '--list-assets',
        action='store_true',
        help='Liste alle Assets auf (zum Debuggen)'
    )
    
    args = parser.parse_args()
    
    print("=" * 70)
    print("🚀 Thingsboard Asset Import")
    print("=" * 70)
    
    if args.dry_run:
        print("⚠️  DRY-RUN MODUS: Es werden keine Assets erstellt!")
    else:
        # Login zu Thingsboard
        token = login_to_thingsboard()
        if not token:
            print("❌ Konnte nicht zu Thingsboard einloggen")
            sys.exit(1)
    
    # Wenn --list-assets, liste alle Assets auf
    if args.list_assets and not args.dry_run:
        print("\n📋 Liste aller Assets:")
        url = f"{THINGSBOARD_BASE_URL}/api/tenant/assets"
        headers = {
            'X-Authorization': f'Bearer {token}',
            'Content-Type': 'application/json'
        }
        
        page = 0
        page_size = 100
        has_next = True
        all_assets = []
        
        while has_next:
            params = {
                'page': page,
                'pageSize': page_size
            }
            
            response = requests.get(url, headers=headers, params=params, timeout=30)
            
            if response.status_code == 200:
                data = response.json()
                assets = data.get('data', [])
                all_assets.extend(assets)
                
                has_next = data.get('hasNext', False)
                page += 1
                
                if not has_next:
                    break
            else:
                print(f"❌ Fehler beim Abrufen der Assets: {response.status_code}")
                break
        
        print(f"   Gefunden: {len(all_assets)} Assets")
        print("\n   Assets mit 'JHRZO' im Namen:")
        jhrzo_assets = [a for a in all_assets if 'JHRZO' in a.get('name', '').upper()]
        for asset in sorted(jhrzo_assets, key=lambda x: x.get('name', '')):
            print(f"     - {asset.get('name')} (ID: {asset.get('id', {}).get('id', 'N/A')})")
        
        if not jhrzo_assets:
            print("     (Keine Assets mit 'JHRZO' im Namen gefunden)")
        
        sys.exit(0)
    
    # Lade Excel-Datei
    df = load_excel_file(args.excel_file)
    if df is None:
        sys.exit(1)
    
    # Verarbeite Assets
    df_result = process_assets(df, dry_run=args.dry_run)
    
    if df_result is not None:
        # Speichere Excel-Datei mit Asset-Namen in Spalten P, Q, R
        print(f"\n💾 Speichere Excel-Datei mit Asset-Namen...")
        try:
            # Verwende openpyxl um die Datei zu speichern und Formatierung zu behalten
            from openpyxl import load_workbook
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            # Lade die ursprüngliche Excel-Datei mit openpyxl
            wb = load_workbook(args.excel_file)
            ws = wb.active
            
            # Schreibe die Asset-Namen in Spalten Ebene 1-6 (H, I, J, K, L, M)
            # Finde Spaltenindizes für Ebene-Spalten
            ebene_cols = {}
            for col_idx, col_name in enumerate(ws[1], 1):
                col_name_str = str(col_name.value).strip() if col_name.value else ''
                if col_name_str.startswith('Ebene'):
                    ebene_num = col_name_str.replace('Ebene', '').strip()
                    if ebene_num.isdigit():
                        ebene_cols[int(ebene_num)] = col_idx
            
            for idx, row in df_result.iterrows():
                # Excel: Zeile 1 = Header, Zeile 2 = übersprungen, Zeile 3+ = Daten
                # DataFrame Index 0 entspricht Excel Zeile 3
                row_num = idx + 3  # +3 weil Excel bei Zeile 1 startet, Zeile 1 ist Header, Zeile 2 wird übersprungen
                
                # Ebene 1 (Spalte H, 8. Spalte) - Level 1
                asset_name_level1 = row.get('Ebene 1', '')
                if asset_name_level1 and str(asset_name_level1).strip():
                    col = ebene_cols.get(1, 8)  # Fallback auf Spalte 8
                    ws.cell(row=row_num, column=col, value=str(asset_name_level1))
                
                # Ebene 2 (Spalte I, 9. Spalte) - Level 2
                asset_name_level2 = row.get('Ebene 2', '')
                if asset_name_level2 and str(asset_name_level2).strip():
                    col = ebene_cols.get(2, 9)  # Fallback auf Spalte 9
                    ws.cell(row=row_num, column=col, value=str(asset_name_level2))
                
                # Ebene 3 (Spalte J, 10. Spalte) - Level 3
                asset_name_level3 = row.get('Ebene 3', '')
                if asset_name_level3 and str(asset_name_level3).strip():
                    col = ebene_cols.get(3, 10)  # Fallback auf Spalte 10
                    ws.cell(row=row_num, column=col, value=str(asset_name_level3))
                
                # Ebene 4 (Spalte K, 11. Spalte) - Level 4
                asset_name_level4 = row.get('Ebene 4', '')
                if asset_name_level4 and str(asset_name_level4).strip():
                    col = ebene_cols.get(4, 11)  # Fallback auf Spalte 11
                    ws.cell(row=row_num, column=col, value=str(asset_name_level4))
                
                # Ebene 5 (Spalte L, 12. Spalte) - Level 5
                asset_name_level5 = row.get('Ebene 5', '')
                if asset_name_level5 and str(asset_name_level5).strip():
                    col = ebene_cols.get(5, 12)  # Fallback auf Spalte 12
                    ws.cell(row=row_num, column=col, value=str(asset_name_level5))
                
                # Ebene 6 (Spalte M, 13. Spalte) - Level 6
                asset_name_level6 = row.get('Ebene 6', '')
                if asset_name_level6 and str(asset_name_level6).strip():
                    col = ebene_cols.get(6, 13)  # Fallback auf Spalte 13
                    ws.cell(row=row_num, column=col, value=str(asset_name_level6))
            
            # Speichere die Datei
            wb.save(args.excel_file)
            print(f"✅ Excel-Datei gespeichert: {args.excel_file}")
            
        except ImportError:
            # Fallback: Verwende pandas to_excel wenn openpyxl nicht verfügbar
            print(f"⚠️  openpyxl nicht verfügbar, verwende pandas to_excel...")
            df_result.to_excel(args.excel_file, index=False)
            print(f"✅ Excel-Datei gespeichert: {args.excel_file}")
        except Exception as e:
            print(f"❌ Fehler beim Speichern der Excel-Datei: {e}")
            import traceback
            traceback.print_exc()
        
        print("\n" + "=" * 70)
        print("✅ Import erfolgreich abgeschlossen!")
        print("=" * 70)
        sys.exit(0)
    else:
        print("\n" + "=" * 70)
        print("❌ Import fehlgeschlagen!")
        print("=" * 70)
        sys.exit(1)

if __name__ == "__main__":
    main()

